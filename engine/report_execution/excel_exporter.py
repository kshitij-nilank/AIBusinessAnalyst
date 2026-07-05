"""Excel export abstraction for report execution."""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Protocol

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from engine.report_execution.exceptions import ExcelExportError, ExcelFileWriteError


logger = logging.getLogger(__name__)


class DataFrameExporter(Protocol):
    """Protocol for DataFrame exporters."""

    def export(self, dataframe: pd.DataFrame, output_filename: str) -> str:
        """Export a DataFrame and return the created filename."""


class ExcelExporter:
    """Export report DataFrames to formatted Excel workbooks."""

    def __init__(self, output_dir: str | Path = "reports/output") -> None:
        """Create an exporter that writes files below output_dir."""

        self.output_dir = Path(output_dir)

    def export(self, dataframe: pd.DataFrame, output_filename: str) -> str:
        """Save a DataFrame to Excel with basic workbook formatting."""

        logger.info("Creating Excel...", extra={"event": "excel_export_started"})
        output_path = self.prepare_output_path(output_filename)

        try:
            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                dataframe.to_excel(writer, index=False, sheet_name="Report")
                worksheet = writer.book["Report"]
                self._freeze_top_row(worksheet)
                self._apply_header_style(worksheet)
                self._auto_adjust_column_widths(worksheet)
                self._apply_number_formats(worksheet)
        except PermissionError as exc:
            raise ExcelFileWriteError(f"Cannot write Excel file: {exc}") from exc
        except OSError as exc:
            raise ExcelFileWriteError(f"Cannot save Excel file: {exc}") from exc
        except Exception as exc:
            raise ExcelExportError(f"Excel export failed: {exc}") from exc

        resolved_path = str(output_path.resolve())
        logger.info(
            "Excel Generated Successfully. Saved To: %s",
            resolved_path,
            extra={"event": "excel_export_completed", "output_file": resolved_path},
        )
        return resolved_path

    def prepare_output_path(self, output_filename: str) -> Path:
        """Create the output folder and return a non-colliding output path."""

        try:
            self.output_dir.mkdir(parents=True, exist_ok=True)
            return self._available_output_path(output_filename)
        except PermissionError as exc:
            raise ExcelFileWriteError(f"Cannot create output folder: {exc}") from exc
        except OSError as exc:
            raise ExcelFileWriteError(f"Cannot prepare output folder: {exc}") from exc

    def _available_output_path(self, output_filename: str) -> Path:
        """Return a non-colliding output path for the requested filename."""

        requested_path = self.output_dir / output_filename
        if not requested_path.exists():
            return requested_path

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return requested_path.with_name(
            f"{requested_path.stem}_{timestamp}{requested_path.suffix}"
        )

    @staticmethod
    def _freeze_top_row(worksheet) -> None:
        """Freeze the first row for easier report scanning."""

        worksheet.freeze_panes = "A2"

    @staticmethod
    def _apply_header_style(worksheet) -> None:
        """Apply a simple header style."""

        fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        font = Font(color="FFFFFF", bold=True)
        for cell in worksheet[1]:
            cell.fill = fill
            cell.font = font

    @staticmethod
    def _auto_adjust_column_widths(worksheet) -> None:
        """Resize columns based on visible cell values."""

        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 40)

    @staticmethod
    def _apply_number_formats(worksheet) -> None:
        """Apply basic numeric formatting to numeric cells."""

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, int):
                    cell.number_format = "#,##0"
                elif isinstance(cell.value, float):
                    cell.number_format = "#,##0.00"
